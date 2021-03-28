import argparse
from openpyxl import Workbook, load_workbook
from datetime import datetime

parser = argparse.ArgumentParser()
parser.add_argument("command", help="command for mngr trk")
parser.add_argument("name", help="name of work for tracking")

args = parser.parse_args()

workbook_name = 'table_trk.xlsx'
now = datetime.now()
current_month = now.strftime("%B")
current_time = now.strftime("%H:%M:%S")

def create_workbook():
    headers = ['name', 'start_time', 'stop_time', 'total_time']
    wb = Workbook()
    page = wb.active
    page.title = current_month 
    page.append(headers)
    wb.save(workbook_name)


def stop():
    
    try:
        wb_load = load_workbook(workbook_name)        
        page = wb_load.active
        existing_names = []
        
        for i in range(1, len(page['A'])):
            existing_names.append(page['A'][i].value)

        if args.name not in existing_names: print("this name doesnt exist")

        else:
            for i in range(1, len(page['A'])):
                if args.name == page['A'][i].value:
                    previous_time = datetime.strptime(page['B'][i].value, "%H:%M:%S")
                    current_time_obj = datetime.strptime(current_time, "%H:%M:%S")
                    total_time = page['D'][i].value
                
                    page['C'][i].value = current_time
                    if total_time != None:
                        page['D'][i].value = (current_time_obj - previous_time) + total_time
                    else: 
                        page['D'][i].value = current_time_obj - previous_time
                    
        wb_load.save(workbook_name)

    except FileNotFoundError:
        print("File not found =( try create_book command")


def start():    

    try:
        wb_load = load_workbook(workbook_name)        
        page = wb_load.active
        existing_names = []
        
        for i in range(1, len(page['A'])):
            existing_names.append(page['A'][i].value)

        if(args.name not in existing_names):
            page.append([args.name, current_time, '',''])
        else:
            for i in range(1, len(page['A'])):
                if args.name == page['A'][i].value:
                    page['B'][i].value = current_time
                    page['C'][i].value = None
                    
        wb_load.save(workbook_name)

    except FileNotFoundError:
        print("File not found =( try create_book command")


if args.command == "create_book":
    create_workbook()
if args.command == "start":
    start()
if args.command == "stop":
    stop()
